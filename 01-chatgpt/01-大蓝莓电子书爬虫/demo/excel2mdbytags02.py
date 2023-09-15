import openpyxl
import os

# 创建"md"目录，如果不存在的话
if not os.path.exists("md"):
    os.mkdir("md")

# 打开Excel文件
excel_file = openpyxl.load_workbook('dalanmei_books_by_tags.xlsx')

# 创建总目录README.md
with open('README.md', 'w', encoding='utf-8') as readme_file:
    readme_file.write('# 目录\n\n')

    # 创建字典来存储每个sheet的数据行数
    sheet_counts = {}

    # 遍历Excel文件的每个工作表
    for sheet_name in excel_file.sheetnames:
        sheet = excel_file[sheet_name]

        # 获取工作表的列名（表头）
        column_names = [cell.value for cell in sheet[1]]

        # 确定所需列的索引
        column_indices = {}
        for i, column in enumerate(column_names):
            if column in ["书名", "EPUB下载链接", "MOBI下载链接", "AZW3下载链接"]:
                column_indices[column] = i + 1

        # 获取工作表的数据行数
        data_rows = list(sheet.iter_rows(min_row=2, values_only=True))
        sheet_counts[sheet_name] = len(data_rows)

        # 创建Markdown文件，文件名与工作表名称相同，保存在"md"目录下
        md_directory = 'md'
        md_filename = os.path.join(md_directory, f'{sheet_name}.md')
        with open(md_filename, 'w', encoding='utf-8') as md_file:
            # 写入工作表名称作为Markdown的一级标题
            md_file.write(f'# {sheet_name}\n\n')

            # 写入表头作为Markdown的表头
            md_file.write('| 书名 | epub | mobi | azw3 |\n')
            md_file.write('| --- | --- | --- | --- |\n')

            # 遍历每一行数据并写入Markdown文件
            for row in data_rows:
                row_data = [
                    str(row[i]) if column_names[i]
                    in ["书名", "EPUB下载链接", "MOBI下载链接", "AZW3下载链接"] else ''
                    for i in range(len(column_names))
                ]
                _, book_name, _, _, _, _, epub_link, mobi_link, azw3_link = row_data

                # 使用"书名"作为默认链接
                md_file.write(
                    f'| [{book_name}]({epub_link}) | [{book_name}]({mobi_link}) | [{book_name}]({azw3_link}) |\n'
                )

        print(f"Markdown文件 '{md_filename}' 已生成。")

    # 写入总目录README.md
    for sheet_name, count in sheet_counts.items():
        readme_file.write(f'- {sheet_name}({count})\n')

print("Markdown文件和总目录README.md已成功生成。")
