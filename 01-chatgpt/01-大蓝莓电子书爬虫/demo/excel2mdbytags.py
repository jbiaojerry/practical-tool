import openpyxl
import os

# 打开Excel文件
excel_file = openpyxl.load_workbook('dalanmei_books_by_tags.xlsx')

# 创建总目录README.md
with open('README.md', 'w', encoding='utf-8') as readme_file:
    readme_file.write('# 目录\n\n')

    # 遍历Excel文件的每个工作表
    for sheet_name in excel_file.sheetnames:
        sheet = excel_file[sheet_name]

        # 创建Markdown文件，文件名与工作表名称相同
        md_filename = f'{sheet_name}.md'
        with open(md_filename, 'w', encoding='utf-8') as md_file:
            # 写入工作表名称作为Markdown的一级标题
            md_file.write(f'# {sheet_name}\n\n')

            # 获取工作表的列名（表头）
            column_names = [cell.value for cell in sheet[1]]

            # 写入表头作为Markdown的表头
            md_file.write('| ' + ' | '.join(column_names) + ' |\n')
            md_file.write('| ' + ' | '.join(['---' for _ in column_names]) + ' |\n')

            # 遍历每一行数据并写入Markdown文件
            for row in sheet.iter_rows(min_row=2, values_only=True):
                row_data = [str(cell) if cell is not None else '' for cell in row]
                md_file.write('| ' + ' | '.join(row_data) + ' |\n')

        # 在总目录README.md中添加链接到Markdown文件
        readme_file.write(f'- [{sheet_name}]({md_filename})\n')

print("Markdown文件和总目录README.md已成功生成。")
