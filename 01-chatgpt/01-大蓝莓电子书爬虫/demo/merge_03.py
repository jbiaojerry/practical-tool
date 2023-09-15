import openpyxl

# 打开本地的a.xlsx文件
a_workbook = openpyxl.load_workbook('fetch_books_async_01.xlsx')

# 创建一个字典，用于存储ID对应的EPUB下载链接和URL
id_data = {}

# 遍历a.xlsx中的所有Sheet
for a_sheet in a_workbook.sheetnames:
    sheet = a_workbook[a_sheet]
    for row in sheet.iter_rows(min_row=2, values_only=True):
        if len(row) >= 6:
            id_field, url, _, _, epub_link, *_ = row  # 仅提取所需的字段，忽略多余的字段
            if id_field is not None:
                # 将id_field转换为整数类型
                id_field = int(id_field)
                id_data[id_field] = {'epub_link': epub_link, 'url': url}

# 打开books.xlsx文件
books_workbook = openpyxl.load_workbook('books_async_11.xlsx')

# 遍历books.xlsx中的所有Sheet
for books_sheet in books_workbook.sheetnames:
    sheet = books_workbook[books_sheet]
    for row in sheet.iter_rows(min_row=1, max_row=1, values_only=True):
        if 'EPUB下载链接' not in row:
            # 如果标题行中没有'EPUB下载链接'和'URL'这两列，则添加它们
            sheet.insert_cols(6, amount=2)  # 在第6列之后插入2列
            sheet.cell(row=1, column=6, value='EPUB下载链接')  # 设置第6列的标题为'EPUB下载链接'
            sheet.cell(row=1, column=7, value='URL')  # 设置第7列的标题为'URL'

# 遍历books.xlsx中的所有Sheet
for books_sheet in books_workbook.sheetnames:
    sheet = books_workbook[books_sheet]
    for row in sheet.iter_rows(min_row=2, values_only=True):
        if len(row) >= 1:
            id_field, *_ = row  # 仅提取所需的字段，忽略多余的字段
            if id_field is not None:
                id_field = int(id_field)  # 将id_field转换为整数类型
                data = id_data.get(id_field, None)
                if data:
                    epub_link, url = data['epub_link'], data['url']
                    # 找到books.xlsx中的相应行并添加EPUB下载链接和URL
                    sheet.append(['', '', '', '', '', epub_link, url])  # 使用append在末尾添加新数据

# 保存更新后的books.xlsx文件
books_workbook.save('books_async_03_updated.xlsx')

print("已添加'EPUB下载链接'和'URL'两列，并保存到 books_async_11_updated.xlsx 文件。")
