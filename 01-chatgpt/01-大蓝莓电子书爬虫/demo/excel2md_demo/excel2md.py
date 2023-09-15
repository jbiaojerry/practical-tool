import openpyxl

# 打开Excel文件
excel_file = openpyxl.load_workbook('dalanmei_books_by_tags.xlsx')

# 创建Markdown文件
markdown_file = open('README.md', 'w', encoding='utf-8')

# 遍历Excel文件的每个工作表
for sheet_name in excel_file.sheetnames:
    sheet = excel_file[sheet_name]

    # 写入工作表名称作为Markdown的一级标题
    markdown_file.write(f'# {sheet_name}\n')

    # 获取工作表的列名（表头）
    column_names = [cell.value for cell in sheet[1]]

    # 写入表头作为Markdown的表头
    markdown_file.write('| ' + ' | '.join(column_names) + ' |\n')
    markdown_file.write('| ' + ' | '.join(['---' for _ in column_names]) + ' |\n')

    # 遍历每一行数据并写入Markdown文件
    for row in sheet.iter_rows(min_row=2, values_only=True):
        row_data = [str(cell) if cell is not None else '' for cell in row]
        markdown_file.write('| ' + ' | '.join(row_data) + ' |\n')

# 关闭Markdown文件
markdown_file.close()

print("Excel文件已成功转换为Markdown格式。")
