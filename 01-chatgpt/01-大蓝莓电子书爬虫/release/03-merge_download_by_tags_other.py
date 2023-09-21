import openpyxl

# 打开本地的01-fetch_books_download_url.xlsx文件
a_workbook = openpyxl.load_workbook('01-fetch_books_download_url.xlsx')

# 创建一个字典，用于存储ID对应的EPUB下载链接、URL和书名
id_data = {}

# 遍历a.xlsx中的所有Sheet
for a_sheet in a_workbook.sheetnames:
    sheet = a_workbook[a_sheet]
    for row in sheet.iter_rows(min_row=2, values_only=True):
        if len(row) >= 6:  # 仅提取所需的字段，包括书名
            id_field, book_title, download_url, mobi_link, epub_link, azw3_link, *_ = row
            if id_field is not None:
                # 将id_field转换为整数类型
                id_field = int(id_field)
                
                # 从book_title中提取书名
                book_title = book_title.split('_')[-1]
                id_data[id_field] = {
                    'download_url': download_url,
                    'epub_link': epub_link,
                    'mobi_link': mobi_link,
                    'azw3_link': azw3_link,
                    'book_title': book_title,
                }

# 打开books.xlsx文件
books_workbook = openpyxl.load_workbook('02-fetch_books_by_tags.xlsx')

# 遍历books.xlsx中的所有Sheet
for books_sheet in books_workbook.sheetnames:
    sheet = books_workbook[books_sheet]
    for row in sheet.iter_rows(min_row=1, max_row=1, values_only=True):
        if 'EPUB下载链接' not in row:
            # 如果标题行中没有'EPUB下载链接'和'URL'这两列，则添加它们
            sheet.insert_cols(6, amount=5)  # 在第6列之后插入5列
            sheet.cell(row=1, column=6, value='下载页面')  # 设置第6列的标题为'下载页面'
            sheet.cell(row=1, column=7,
                       value='EPUB下载链接')  # 设置第7列的标题为'EPUB下载链接'
            sheet.cell(row=1, column=8,
                       value='MOBI下载链接')  # 设置第8列的标题为'MOBI下载链接'
            sheet.cell(row=1, column=9,
                       value='AZW3下载链接')  # 设置第9列的标题为'AZW3下载链接'

books_id_data = {}      
# 遍历books.xlsx中的所有Sheet
for books_sheet in books_workbook.sheetnames:
    sheet = books_workbook[books_sheet]
    for index, row in enumerate(sheet.iter_rows(min_row=2, values_only=True),
                                start=2):
        if len(row) >= 1:
            id_field, book_title, *_ = row  # 仅提取所需的字段，忽略多余的字段
            if id_field is not None:
                id_field = int(id_field)  # 将id_field转换为整数类型
                books_id_data[id_field] = {'book_title': book_title,}
                data = id_data.get(id_field, None)
                if data:
                    download_url, epub_link, mobi_link, azw3_link = data[
                        'download_url'], data['epub_link'], data[
                            'mobi_link'], data['azw3_link'],
                    # 找到books.xlsx中的相应行并添加EPUB下载链接和URL
                    sheet.cell(row=index, column=6, value=download_url)
                    sheet.cell(row=index, column=7, value=epub_link)
                    sheet.cell(row=index, column=8, value=mobi_link)
                    sheet.cell(row=index, column=9, value=azw3_link)

# 创建'other'工作表
other_sheet = books_workbook['other'] if 'other' in books_workbook.sheetnames else books_workbook.create_sheet(title='other')
other_sheet.append(['ID', '书名', '下载页面', 'EPUB下载链接', 'MOBI下载链接', 'AZW3下载链接'])

# 查找在'01-fetch_books_download_url.xlsx'中但在'02-fetch_books_by_tags.xlsx'中找不到的ID，并将其添加到'other'工作表
for id_field, data in id_data.items():
    if id_field > 13000:
        print("id", id_field)
    tmp = books_id_data.get(int(id_field), None)
    if not tmp:
        print("未找到:", id_field, data['book_title'])
        other_sheet.append([id_field, data['book_title'], data['download_url'], data['epub_link'], data['mobi_link'], data['azw3_link']])

# 删除原始工作表中无数据的行
sheet = other_sheet
rows_to_delete = []
for index, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
    if len(row) >= 1:
        id_field, *_ = row
        if id_field is not None:
            id_field = int(id_field)
            data = id_data.get(id_field, None)
            if data and all(data.values()):
                continue
    rows_to_delete.append(index)
for row_index in reversed(rows_to_delete):
    sheet.delete_rows(row_index, 1)

# 保存更新后的books.xlsx文件
books_workbook.save('03-dalanmei_books_by_tags_other.xlsx')

print("已添加'EPUB下载链接'、'URL'和'书名'三列，并保存到 03-dalanmei_books_by_tags.xlsx 文件。")
